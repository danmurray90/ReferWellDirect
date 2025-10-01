"""
Analytics and reporting service for referrals and appointments.
"""
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from django.db.models import Avg, Count, F, Max, Min, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
from django.utils import timezone

from .models import Appointment, Candidate, Message, Referral, Task


class AnalyticsService:
    """
    Service for generating analytics and reports.
    """

    def __init__(self) -> None:
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")

    def get_dashboard_metrics(
        self, user: Any, date_range: str = "30d"
    ) -> dict[str, Any]:
        """
        Get key metrics for the dashboard.

        Args:
            user: Current user for permission filtering
            date_range: Date range ('7d', '30d', '90d', '1y')

        Returns:
            Dictionary of dashboard metrics
        """
        try:
            # Calculate date range
            end_date = timezone.now()
            start_date = self._get_start_date(end_date, date_range)

            # Get base querysets
            referrals_qs = self._get_referrals_queryset(user, start_date, end_date)
            appointments_qs = self._get_appointments_queryset(
                user, start_date, end_date
            )
            candidates_qs = self._get_candidates_queryset(user, start_date, end_date)

            # Basic counts
            total_referrals = referrals_qs.count()
            total_appointments = appointments_qs.count()
            total_candidates = candidates_qs.count()

            # Referral metrics
            referral_metrics = self._get_referral_metrics(referrals_qs)

            # Appointment metrics
            appointment_metrics = self._get_appointment_metrics(appointments_qs)

            # Candidate metrics
            candidate_metrics = self._get_candidate_metrics(candidates_qs)

            # Performance metrics
            performance_metrics = self._get_performance_metrics(
                referrals_qs, appointments_qs
            )

            # Trends
            trends = self._get_trends(user, start_date, end_date)

            return {
                "date_range": date_range,
                "start_date": start_date,
                "end_date": end_date,
                "total_referrals": total_referrals,
                "total_appointments": total_appointments,
                "total_candidates": total_candidates,
                "referral_metrics": referral_metrics,
                "appointment_metrics": appointment_metrics,
                "candidate_metrics": candidate_metrics,
                "performance_metrics": performance_metrics,
                "trends": trends,
            }

        except Exception as e:
            self.logger.error(f"Failed to get dashboard metrics: {e}")
            return {}

    def get_referral_analytics(
        self, user: Any, filters: Optional[dict[str, Any]] = None
    ) -> dict[str, Any]:
        """
        Get detailed referral analytics.

        Args:
            user: Current user for permission filtering
            filters: Optional filters to apply

        Returns:
            Dictionary of referral analytics
        """
        try:
            # Get base queryset
            queryset = self._get_referrals_queryset(user)

            # Apply filters
            if filters:
                queryset = self._apply_filters(queryset, filters)

            # Status distribution
            status_distribution = (
                queryset.values("status").annotate(count=Count("id")).order_by("status")
            )

            # Priority distribution
            priority_distribution = (
                queryset.values("priority")
                .annotate(count=Count("id"))
                .order_by("priority")
            )

            # Service type distribution
            service_type_distribution = (
                queryset.values("service_type")
                .annotate(count=Count("id"))
                .order_by("service_type")
            )

            # Modality distribution
            modality_distribution = (
                queryset.values("modality")
                .annotate(count=Count("id"))
                .order_by("modality")
            )

            # Age group distribution
            age_group_distribution = (
                queryset.exclude(patient_age_group__isnull=True)
                .values("patient_age_group")
                .annotate(count=Count("id"))
                .order_by("patient_age_group")
            )

            # Language distribution
            language_distribution = (
                queryset.values("preferred_language")
                .annotate(count=Count("id"))
                .order_by("preferred_language")
            )

            # Processing time analysis
            processing_times = (
                queryset.filter(submitted_at__isnull=False, completed_at__isnull=False)
                .annotate(processing_time=F("completed_at") - F("submitted_at"))
                .aggregate(
                    avg_processing_time=Avg("processing_time"),
                    min_processing_time=Min("processing_time"),
                    max_processing_time=Max("processing_time"),
                )
            )

            # Monthly trends
            monthly_trends = (
                queryset.annotate(month=TruncMonth("created_at"))
                .values("month")
                .annotate(count=Count("id"))
                .order_by("month")
            )

            # Top referrers
            top_referrers = (
                queryset.values("referrer__first_name", "referrer__last_name")
                .annotate(count=Count("id"))
                .order_by("-count")[:10]
            )

            # Top specialisms
            specialism_counts = {}
            for referral in queryset.filter(required_specialisms__isnull=False):
                for specialism in referral.required_specialisms:
                    specialism_counts[specialism] = (
                        specialism_counts.get(specialism, 0) + 1
                    )

            top_specialisms = sorted(
                specialism_counts.items(), key=lambda x: x[1], reverse=True
            )[:10]

            return {
                "status_distribution": list(status_distribution),
                "priority_distribution": list(priority_distribution),
                "service_type_distribution": list(service_type_distribution),
                "modality_distribution": list(modality_distribution),
                "age_group_distribution": list(age_group_distribution),
                "language_distribution": list(language_distribution),
                "processing_times": processing_times,
                "monthly_trends": list(monthly_trends),
                "top_referrers": list(top_referrers),
                "top_specialisms": top_specialisms,
            }

        except Exception as e:
            self.logger.error(f"Failed to get referral analytics: {e}")
            return {}

    def get_appointment_analytics(
        self, user: Any, filters: Optional[dict[str, Any]] = None
    ) -> dict[str, Any]:
        """
        Get detailed appointment analytics.

        Args:
            user: Current user for permission filtering
            filters: Optional filters to apply

        Returns:
            Dictionary of appointment analytics
        """
        try:
            # Get base queryset
            queryset = self._get_appointments_queryset(user)

            # Apply filters
            if filters:
                queryset = self._apply_appointment_filters(queryset, filters)

            # Status distribution
            status_distribution = (
                queryset.values("status").annotate(count=Count("id")).order_by("status")
            )

            # Modality distribution
            modality_distribution = (
                queryset.values("modality")
                .annotate(count=Count("id"))
                .order_by("modality")
            )

            # Duration analysis
            duration_analysis = queryset.aggregate(
                avg_duration=Avg("duration_minutes"),
                min_duration=Min("duration_minutes"),
                max_duration=Max("duration_minutes"),
            )

            # Monthly trends
            monthly_trends = (
                queryset.annotate(month=TruncMonth("scheduled_at"))
                .values("month")
                .annotate(count=Count("id"))
                .order_by("month")
            )

            # Top psychologists
            top_psychologists = (
                queryset.values("psychologist__first_name", "psychologist__last_name")
                .annotate(count=Count("id"))
                .order_by("-count")[:10]
            )

            # Completion rates
            completion_rates = queryset.aggregate(
                total_appointments=Count("id"),
                completed_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.COMPLETED)
                ),
                cancelled_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.CANCELLED)
                ),
                no_show_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.NO_SHOW)
                ),
            )

            # Calculate percentages
            total = completion_rates["total_appointments"]
            if total > 0:
                completion_rates["completion_rate"] = (
                    completion_rates["completed_appointments"] / total
                ) * 100
                completion_rates["cancellation_rate"] = (
                    completion_rates["cancelled_appointments"] / total
                ) * 100
                completion_rates["no_show_rate"] = (
                    completion_rates["no_show_appointments"] / total
                ) * 100
            else:
                completion_rates["completion_rate"] = 0
                completion_rates["cancellation_rate"] = 0
                completion_rates["no_show_rate"] = 0

            return {
                "status_distribution": list(status_distribution),
                "modality_distribution": list(modality_distribution),
                "duration_analysis": duration_analysis,
                "monthly_trends": list(monthly_trends),
                "top_psychologists": list(top_psychologists),
                "completion_rates": completion_rates,
            }

        except Exception as e:
            self.logger.error(f"Failed to get appointment analytics: {e}")
            return {}

    def get_performance_metrics(
        self, user: Any, filters: Optional[dict[str, Any]] = None
    ) -> dict[str, Any]:
        """
        Get performance metrics and KPIs.

        Args:
            user: Current user for permission filtering
            filters: Optional filters to apply

        Returns:
            Dictionary of performance metrics
        """
        try:
            # Get base querysets
            referrals_qs = self._get_referrals_queryset(user)
            appointments_qs = self._get_appointments_queryset(user)
            candidates_qs = self._get_candidates_queryset(user)

            # Apply filters
            if filters:
                referrals_qs = self._apply_filters(referrals_qs, filters)
                appointments_qs = self._apply_appointment_filters(
                    appointments_qs, filters
                )
                candidates_qs = self._apply_candidate_filters(candidates_qs, filters)

            # Referral processing metrics
            referral_metrics = referrals_qs.aggregate(
                total_referrals=Count("id"),
                submitted_referrals=Count(
                    "id",
                    filter=Q(
                        status__in=[
                            Referral.Status.SUBMITTED,
                            Referral.Status.MATCHING,
                            Referral.Status.SHORTLISTED,
                            Referral.Status.HIGH_TOUCH_QUEUE,
                            Referral.Status.COMPLETED,
                        ]
                    ),
                ),
                completed_referrals=Count(
                    "id", filter=Q(status=Referral.Status.COMPLETED)
                ),
                avg_processing_time=Avg(
                    F("completed_at") - F("submitted_at"),
                    filter=Q(submitted_at__isnull=False, completed_at__isnull=False),
                ),
            )

            # Candidate matching metrics
            candidate_metrics = candidates_qs.aggregate(
                total_candidates=Count("id"),
                accepted_candidates=Count(
                    "id", filter=Q(status=Candidate.Status.ACCEPTED)
                ),
                declined_candidates=Count(
                    "id", filter=Q(status=Candidate.Status.DECLINED)
                ),
                avg_score=Avg("final_score"),
                avg_confidence=Avg("confidence_score"),
            )

            # Appointment metrics
            appointment_metrics = appointments_qs.aggregate(
                total_appointments=Count("id"),
                completed_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.COMPLETED)
                ),
                cancelled_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.CANCELLED)
                ),
                no_show_appointments=Count(
                    "id", filter=Q(status=Appointment.Status.NO_SHOW)
                ),
                avg_duration=Avg("duration_minutes"),
            )

            # Calculate rates
            total_referrals = referral_metrics["total_referrals"]
            if total_referrals > 0:
                referral_metrics["submission_rate"] = (
                    referral_metrics["submitted_referrals"] / total_referrals
                ) * 100
                referral_metrics["completion_rate"] = (
                    referral_metrics["completed_referrals"] / total_referrals
                ) * 100
            else:
                referral_metrics["submission_rate"] = 0
                referral_metrics["completion_rate"] = 0

            total_candidates = candidate_metrics["total_candidates"]
            if total_candidates > 0:
                candidate_metrics["acceptance_rate"] = (
                    candidate_metrics["accepted_candidates"] / total_candidates
                ) * 100
                candidate_metrics["decline_rate"] = (
                    candidate_metrics["declined_candidates"] / total_candidates
                ) * 100
            else:
                candidate_metrics["acceptance_rate"] = 0
                candidate_metrics["decline_rate"] = 0

            total_appointments = appointment_metrics["total_appointments"]
            if total_appointments > 0:
                appointment_metrics["completion_rate"] = (
                    appointment_metrics["completed_appointments"] / total_appointments
                ) * 100
                appointment_metrics["cancellation_rate"] = (
                    appointment_metrics["cancelled_appointments"] / total_appointments
                ) * 100
                appointment_metrics["no_show_rate"] = (
                    appointment_metrics["no_show_appointments"] / total_appointments
                ) * 100
            else:
                appointment_metrics["completion_rate"] = 0
                appointment_metrics["cancellation_rate"] = 0
                appointment_metrics["no_show_rate"] = 0

            return {
                "referral_metrics": referral_metrics,
                "candidate_metrics": candidate_metrics,
                "appointment_metrics": appointment_metrics,
            }

        except Exception as e:
            self.logger.error(f"Failed to get performance metrics: {e}")
            return {}

    def generate_report(
        self,
        user: Any,
        report_type: str,
        filters: Optional[dict[str, Any]] = None,
        format: str = "json",
    ) -> dict[str, Any]:
        """
        Generate a comprehensive report.

        Args:
            user: Current user for permission filtering
            report_type: Type of report ('referrals', 'appointments', 'performance', 'comprehensive')
            filters: Optional filters to apply
            format: Report format ('json', 'csv', 'xlsx')

        Returns:
            Dictionary with report data
        """
        try:
            if report_type == "referrals":
                data = self.get_referral_analytics(user, filters)
            elif report_type == "appointments":
                data = self.get_appointment_analytics(user, filters)
            elif report_type == "performance":
                data = self.get_performance_metrics(user, filters)
            elif report_type == "comprehensive":
                data = {
                    "referrals": self.get_referral_analytics(user, filters),
                    "appointments": self.get_appointment_analytics(user, filters),
                    "performance": self.get_performance_metrics(user, filters),
                }
            else:
                return {"error": f"Unknown report type: {report_type}"}

            # Add metadata
            data["report_metadata"] = {
                "generated_at": timezone.now().isoformat(),
                "generated_by": user.get_full_name(),
                "report_type": report_type,
                "filters": filters or {},
            }

            if format == "csv":
                return self._export_csv(data, report_type)
            elif format == "xlsx":
                return self._export_xlsx(data, report_type)
            else:
                return {"success": True, "data": data}

        except Exception as e:
            self.logger.error(f"Failed to generate report: {e}")
            return {"error": str(e)}

    def _get_start_date(self, end_date: Any, date_range: str) -> Any:
        """Get start date based on date range."""
        if date_range == "7d":
            return end_date - timedelta(days=7)
        elif date_range == "30d":
            return end_date - timedelta(days=30)
        elif date_range == "90d":
            return end_date - timedelta(days=90)
        elif date_range == "1y":
            return end_date - timedelta(days=365)
        else:
            return end_date - timedelta(days=30)

    def _get_referrals_queryset(
        self, user: Any, start_date: Any = None, end_date: Any = None
    ) -> Any:
        """Get referrals queryset based on user permissions."""
        queryset = Referral.objects.all()

        if user.is_gp:
            queryset = queryset.filter(referrer=user)
        elif user.is_patient:
            queryset = queryset.filter(patient=user)
        elif not user.is_admin:
            queryset = queryset.none()

        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        return queryset

    def _get_appointments_queryset(
        self, user: Any, start_date: Any = None, end_date: Any = None
    ) -> Any:
        """Get appointments queryset based on user permissions."""
        queryset = Appointment.objects.all()

        if user.is_gp:
            queryset = queryset.filter(referral__referrer=user)
        elif user.is_patient:
            queryset = queryset.filter(patient=user)
        elif user.is_psychologist:
            queryset = queryset.filter(psychologist=user)
        elif not user.is_admin:
            queryset = queryset.none()

        if start_date:
            queryset = queryset.filter(scheduled_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(scheduled_at__lte=end_date)

        return queryset

    def _get_candidates_queryset(
        self, user: Any, start_date: Any = None, end_date: Any = None
    ) -> Any:
        """Get candidates queryset based on user permissions."""
        queryset = Candidate.objects.all()

        if user.is_gp:
            queryset = queryset.filter(referral__referrer=user)
        elif user.is_patient:
            queryset = queryset.filter(referral__patient=user)
        elif user.is_psychologist:
            queryset = queryset.filter(psychologist=user)
        elif not user.is_admin:
            queryset = queryset.none()

        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        return queryset

    def _apply_filters(self, queryset: Any, filters: dict[str, Any]) -> Any:
        """Apply filters to referrals queryset."""
        if filters.get("status"):
            queryset = queryset.filter(status=filters["status"])
        if filters.get("priority"):
            queryset = queryset.filter(priority=filters["priority"])
        if filters.get("service_type"):
            queryset = queryset.filter(service_type=filters["service_type"])
        if filters.get("modality"):
            queryset = queryset.filter(modality=filters["modality"])
        if filters.get("date_from"):
            queryset = queryset.filter(created_at__gte=filters["date_from"])
        if filters.get("date_to"):
            queryset = queryset.filter(created_at__lte=filters["date_to"])

        return queryset

    def _apply_appointment_filters(self, queryset: Any, filters: dict[str, Any]) -> Any:
        """Apply filters to appointments queryset."""
        if filters.get("status"):
            queryset = queryset.filter(status=filters["status"])
        if filters.get("modality"):
            queryset = queryset.filter(modality=filters["modality"])
        if filters.get("date_from"):
            queryset = queryset.filter(scheduled_at__gte=filters["date_from"])
        if filters.get("date_to"):
            queryset = queryset.filter(scheduled_at__lte=filters["date_to"])

        return queryset

    def _apply_candidate_filters(self, queryset: Any, filters: dict[str, Any]) -> Any:
        """Apply filters to candidates queryset."""
        if filters.get("status"):
            queryset = queryset.filter(status=filters["status"])
        if filters.get("date_from"):
            queryset = queryset.filter(created_at__gte=filters["date_from"])
        if filters.get("date_to"):
            queryset = queryset.filter(created_at__lte=filters["date_to"])

        return queryset

    def _get_referral_metrics(self, queryset: Any) -> Any:
        """Get referral-specific metrics."""
        return queryset.aggregate(
            total=Count("id"),
            submitted=Count(
                "id",
                filter=Q(
                    status__in=[
                        Referral.Status.SUBMITTED,
                        Referral.Status.MATCHING,
                        Referral.Status.SHORTLISTED,
                        Referral.Status.HIGH_TOUCH_QUEUE,
                        Referral.Status.COMPLETED,
                    ]
                ),
            ),
            completed=Count("id", filter=Q(status=Referral.Status.COMPLETED)),
            cancelled=Count("id", filter=Q(status=Referral.Status.CANCELLED)),
            rejected=Count("id", filter=Q(status=Referral.Status.REJECTED)),
        )

    def _get_appointment_metrics(self, queryset: Any) -> Any:
        """Get appointment-specific metrics."""
        return queryset.aggregate(
            total=Count("id"),
            scheduled=Count("id", filter=Q(status=Appointment.Status.SCHEDULED)),
            confirmed=Count("id", filter=Q(status=Appointment.Status.CONFIRMED)),
            completed=Count("id", filter=Q(status=Appointment.Status.COMPLETED)),
            cancelled=Count("id", filter=Q(status=Appointment.Status.CANCELLED)),
            no_show=Count("id", filter=Q(status=Appointment.Status.NO_SHOW)),
        )

    def _get_candidate_metrics(self, queryset: Any) -> Any:
        """Get candidate-specific metrics."""
        return queryset.aggregate(
            total=Count("id"),
            pending=Count("id", filter=Q(status=Candidate.Status.PENDING)),
            shortlisted=Count("id", filter=Q(status=Candidate.Status.SHORTLISTED)),
            invited=Count("id", filter=Q(status=Candidate.Status.INVITED)),
            accepted=Count("id", filter=Q(status=Candidate.Status.ACCEPTED)),
            declined=Count("id", filter=Q(status=Candidate.Status.DECLINED)),
            expired=Count("id", filter=Q(status=Candidate.Status.EXPIRED)),
        )

    def _get_performance_metrics(self, referrals_qs: Any, appointments_qs: Any) -> Any:
        """Get performance metrics."""
        return {
            "avg_processing_time": referrals_qs.filter(
                submitted_at__isnull=False, completed_at__isnull=False
            ).aggregate(avg_time=Avg(F("completed_at") - F("submitted_at")))[
                "avg_time"
            ],
            "avg_appointment_duration": appointments_qs.aggregate(
                avg_duration=Avg("duration_minutes")
            )["avg_duration"],
            "total_revenue": 0,  # Placeholder for future revenue tracking
        }

    def _get_trends(self, user: Any, start_date: Any, end_date: Any) -> Any:
        """Get trend data."""
        # Daily referral trends
        daily_referrals = (
            self._get_referrals_queryset(user, start_date, end_date)
            .annotate(date=TruncDate("created_at"))
            .values("date")
            .annotate(count=Count("id"))
            .order_by("date")
        )

        # Daily appointment trends
        daily_appointments = (
            self._get_appointments_queryset(user, start_date, end_date)
            .annotate(date=TruncDate("scheduled_at"))
            .values("date")
            .annotate(count=Count("id"))
            .order_by("date")
        )

        return {
            "daily_referrals": list(daily_referrals),
            "daily_appointments": list(daily_appointments),
        }

    def _export_csv(self, data: Any, report_type: str) -> Any:
        """Export data as CSV."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # This is a simplified CSV export - in practice, you'd want more sophisticated formatting
        writer.writerow(["Report Type", report_type])
        writer.writerow(
            ["Generated At", data.get("report_metadata", {}).get("generated_at", "")]
        )
        writer.writerow([])

        # Add data rows based on report type
        if report_type == "referrals":
            writer.writerow(["Metric", "Value"])
            for key, value in data.items():
                if key != "report_metadata":
                    writer.writerow([key, value])

        return {
            "success": True,
            "data": output.getvalue(),
            "content_type": "text/csv",
            "filename": f'{report_type}_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv',
        }

    def _export_xlsx(self, data: Any, report_type: str) -> Any:
        """Export data as XLSX."""
        try:
            import openpyxl  # type: ignore[import]
            from openpyxl.styles import Alignment, Font  # type: ignore[import]

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"{report_type.title()} Report"

            # Add metadata
            worksheet.cell(row=1, column=1, value="Report Type")
            worksheet.cell(row=1, column=2, value=report_type)
            worksheet.cell(row=2, column=1, value="Generated At")
            worksheet.cell(
                row=2,
                column=2,
                value=data.get("report_metadata", {}).get("generated_at", ""),
            )

            # Add data (simplified - in practice, you'd want more sophisticated formatting)
            row = 4
            for key, value in data.items():
                if key != "report_metadata":
                    worksheet.cell(row=row, column=1, value=key)
                    worksheet.cell(row=row, column=2, value=str(value))
                    row += 1

            # Save to bytes
            import io

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            return {
                "success": True,
                "data": output.getvalue(),
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filename": f'{report_type}_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            }

        except ImportError:
            return {
                "success": False,
                "error": "openpyxl is required for XLSX export. Install with: pip install openpyxl",
            }
