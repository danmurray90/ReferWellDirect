"""
Advanced search and filtering service for referrals.
"""
import logging
from typing import Dict, List, Any, Optional, Tuple
from django.db.models import Q, F, Count, Avg, Max, Min
from django.contrib.postgres.search import SearchVector, SearchQuery, SearchRank
from django.contrib.postgres.aggregates import StringAgg
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Referral, Candidate, Appointment, Message, Task


class AdvancedSearchService:
    """
    Advanced search and filtering service for referrals.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    def search_referrals(
        self,
        user,
        search_params: Dict[str, Any],
        page: int = 1,
        page_size: int = 20
    ) -> Tuple[List[Referral], Dict[str, Any]]:
        """
        Perform advanced search on referrals.
        
        Args:
            user: Current user for permission filtering
            search_params: Dictionary of search parameters
            page: Page number for pagination
            page_size: Number of items per page
            
        Returns:
            Tuple of (referrals, metadata)
        """
        try:
            # Start with base queryset based on user permissions
            queryset = self._get_base_queryset(user)
            
            # Apply search filters
            queryset = self._apply_search_filters(queryset, search_params)
            
            # Apply sorting
            queryset = self._apply_sorting(queryset, search_params.get('sort', 'created_at'))
            
            # Get total count before pagination
            total_count = queryset.count()
            
            # Apply pagination
            start = (page - 1) * page_size
            end = start + page_size
            referrals = list(queryset[start:end])
            
            # Calculate pagination metadata
            total_pages = (total_count + page_size - 1) // page_size
            has_next = page < total_pages
            has_previous = page > 1
            
            metadata = {
                'total_count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': total_pages,
                'has_next': has_next,
                'has_previous': has_previous,
                'next_page': page + 1 if has_next else None,
                'previous_page': page - 1 if has_previous else None,
            }
            
            return referrals, metadata
            
        except Exception as e:
            self.logger.error(f"Search failed: {e}")
            return [], {'total_count': 0, 'page': 1, 'page_size': page_size, 'total_pages': 0, 'has_next': False, 'has_previous': False}
    
    def _get_base_queryset(self, user):
        """Get base queryset based on user permissions."""
        if user.is_gp:
            return Referral.objects.filter(referrer=user).select_related('patient', 'referrer')
        elif user.is_patient:
            return Referral.objects.filter(patient=user).select_related('patient', 'referrer')
        elif user.is_admin:
            return Referral.objects.all().select_related('patient', 'referrer')
        else:
            return Referral.objects.none()
    
    def _apply_search_filters(self, queryset, search_params):
        """Apply all search filters to the queryset."""
        # Text search
        if search_params.get('q'):
            queryset = self._apply_text_search(queryset, search_params['q'])
        
        # Status filter
        if search_params.get('status'):
            queryset = queryset.filter(status=search_params['status'])
        
        # Priority filter
        if search_params.get('priority'):
            queryset = queryset.filter(priority=search_params['priority'])
        
        # Service type filter
        if search_params.get('service_type'):
            queryset = queryset.filter(service_type=search_params['service_type'])
        
        # Modality filter
        if search_params.get('modality'):
            queryset = queryset.filter(modality=search_params['modality'])
        
        # Date range filters
        if search_params.get('created_from'):
            queryset = queryset.filter(created_at__gte=search_params['created_from'])
        
        if search_params.get('created_to'):
            queryset = queryset.filter(created_at__lte=search_params['created_to'])
        
        if search_params.get('submitted_from'):
            queryset = queryset.filter(submitted_at__gte=search_params['submitted_from'])
        
        if search_params.get('submitted_to'):
            queryset = queryset.filter(submitted_at__lte=search_params['submitted_to'])
        
        # Patient age group filter
        if search_params.get('patient_age_group'):
            queryset = queryset.filter(patient_age_group=search_params['patient_age_group'])
        
        # Language filter
        if search_params.get('preferred_language'):
            queryset = queryset.filter(preferred_language=search_params['preferred_language'])
        
        # Specialism filter
        if search_params.get('required_specialisms'):
            specialisms = search_params['required_specialisms']
            if isinstance(specialisms, str):
                specialisms = [specialisms]
            for specialism in specialisms:
                queryset = queryset.filter(required_specialisms__contains=[specialism])
        
        # Geographic filters
        if search_params.get('max_distance_km'):
            queryset = queryset.filter(max_distance_km__lte=search_params['max_distance_km'])
        
        # Has candidates filter
        if search_params.get('has_candidates') is not None:
            if search_params['has_candidates']:
                queryset = queryset.filter(candidates__isnull=False).distinct()
            else:
                queryset = queryset.filter(candidates__isnull=True)
        
        # Has appointments filter
        if search_params.get('has_appointments') is not None:
            if search_params['has_appointments']:
                queryset = queryset.filter(appointments__isnull=False).distinct()
            else:
                queryset = queryset.filter(appointments__isnull=True)
        
        # Candidate count filters
        if search_params.get('min_candidates'):
            queryset = queryset.annotate(candidate_count=Count('candidates')).filter(
                candidate_count__gte=search_params['min_candidates']
            )
        
        if search_params.get('max_candidates'):
            queryset = queryset.annotate(candidate_count=Count('candidates')).filter(
                candidate_count__lte=search_params['max_candidates']
            )
        
        # Score filters
        if search_params.get('min_score'):
            queryset = queryset.annotate(max_score=Max('candidates__final_score')).filter(
                max_score__gte=search_params['min_score']
            )
        
        if search_params.get('max_score'):
            queryset = queryset.annotate(max_score=Max('candidates__final_score')).filter(
                max_score__lte=search_params['max_score']
            )
        
        return queryset
    
    def _apply_text_search(self, queryset, query):
        """Apply full-text search to the queryset."""
        if not query.strip():
            return queryset
        
        # Create search vector for relevant fields
        search_vector = SearchVector(
            'referral_id',
            'presenting_problem',
            'condition_description',
            'clinical_notes',
            'urgency_notes',
            'patient__first_name',
            'patient__last_name',
            'referrer__first_name',
            'referrer__last_name'
        )
        
        # Create search query
        search_query = SearchQuery(query)
        
        # Apply search with ranking
        queryset = queryset.annotate(
            search_rank=SearchRank(search_vector, search_query)
        ).filter(search_rank__gte=0.1).order_by('-search_rank')
        
        return queryset
    
    def _apply_sorting(self, queryset, sort_field):
        """Apply sorting to the queryset."""
        sort_mapping = {
            'created_at': '-created_at',
            'created_at_asc': 'created_at',
            'submitted_at': '-submitted_at',
            'submitted_at_asc': 'submitted_at',
            'priority': '-priority',
            'priority_asc': 'priority',
            'status': 'status',
            'status_desc': '-status',
            'patient_name': 'patient__last_name',
            'patient_name_desc': '-patient__last_name',
            'referrer_name': 'referrer__last_name',
            'referrer_name_desc': '-referrer__last_name',
            'score': '-candidates__final_score',
            'score_asc': 'candidates__final_score',
        }
        
        sort_order = sort_mapping.get(sort_field, '-created_at')
        
        if 'candidates__final_score' in sort_order:
            # For score sorting, we need to handle the case where there might be multiple candidates
            queryset = queryset.annotate(
                max_score=Max('candidates__final_score')
            ).order_by(sort_order.replace('candidates__final_score', 'max_score'))
        else:
            queryset = queryset.order_by(sort_order)
        
        return queryset
    
    def get_search_facets(self, user, search_params: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Get search facets for filtering UI.
        
        Args:
            user: Current user for permission filtering
            search_params: Current search parameters
            
        Returns:
            Dictionary of facets with counts
        """
        try:
            # Get base queryset
            queryset = self._get_base_queryset(user)
            
            # Apply current filters except the one we're getting facets for
            current_facets = ['status', 'priority', 'service_type', 'modality', 'patient_age_group', 'preferred_language']
            for facet in current_facets:
                if facet in search_params:
                    del search_params[facet]
            
            queryset = self._apply_search_filters(queryset, search_params)
            
            facets = {}
            
            # Status facets
            status_counts = queryset.values('status').annotate(count=Count('id')).order_by('status')
            facets['status'] = [
                {'value': item['status'], 'label': dict(Referral.Status.choices).get(item['status'], item['status']), 'count': item['count']}
                for item in status_counts
            ]
            
            # Priority facets
            priority_counts = queryset.values('priority').annotate(count=Count('id')).order_by('priority')
            facets['priority'] = [
                {'value': item['priority'], 'label': dict(Referral.Priority.choices).get(item['priority'], item['priority']), 'count': item['count']}
                for item in priority_counts
            ]
            
            # Service type facets
            service_type_counts = queryset.values('service_type').annotate(count=Count('id')).order_by('service_type')
            facets['service_type'] = [
                {'value': item['service_type'], 'label': dict(Referral.ServiceType.choices).get(item['service_type'], item['service_type']), 'count': item['count']}
                for item in service_type_counts
            ]
            
            # Modality facets
            modality_counts = queryset.values('modality').annotate(count=Count('id')).order_by('modality')
            facets['modality'] = [
                {'value': item['modality'], 'label': dict(Referral.Modality.choices).get(item['modality'], item['modality']), 'count': item['count']}
                for item in modality_counts
            ]
            
            # Patient age group facets
            age_group_counts = queryset.values('patient_age_group').annotate(count=Count('id')).order_by('patient_age_group')
            facets['patient_age_group'] = [
                {'value': item['patient_age_group'], 'label': item['patient_age_group'] or 'Not specified', 'count': item['count']}
                for item in age_group_counts if item['patient_age_group']
            ]
            
            # Language facets
            language_counts = queryset.values('preferred_language').annotate(count=Count('id')).order_by('preferred_language')
            facets['preferred_language'] = [
                {'value': item['preferred_language'], 'label': item['preferred_language'], 'count': item['count']}
                for item in language_counts
            ]
            
            return facets
            
        except Exception as e:
            self.logger.error(f"Failed to get search facets: {e}")
            return {}
    
    def get_search_suggestions(self, user, query: str, limit: int = 10) -> List[str]:
        """
        Get search suggestions based on query.
        
        Args:
            user: Current user for permission filtering
            query: Search query
            limit: Maximum number of suggestions
            
        Returns:
            List of suggestion strings
        """
        try:
            if not query or len(query) < 2:
                return []
            
            # Get base queryset
            queryset = self._get_base_queryset(user)
            
            # Search in relevant text fields
            suggestions = set()
            
            # Search in presenting problems
            presenting_problems = queryset.filter(
                presenting_problem__icontains=query
            ).values_list('presenting_problem', flat=True)[:limit]
            
            for problem in presenting_problems:
                # Extract relevant parts around the query
                words = problem.lower().split()
                query_words = query.lower().split()
                for i, word in enumerate(words):
                    if any(qw in word for qw in query_words):
                        start = max(0, i - 2)
                        end = min(len(words), i + 3)
                        suggestion = ' '.join(words[start:end])
                        if len(suggestion) > len(query) and len(suggestion) < 100:
                            suggestions.add(suggestion)
            
            # Search in condition descriptions
            conditions = queryset.filter(
                condition_description__icontains=query
            ).values_list('condition_description', flat=True)[:limit]
            
            for condition in conditions:
                if condition:
                    words = condition.lower().split()
                    query_words = query.lower().split()
                    for i, word in enumerate(words):
                        if any(qw in word for qw in query_words):
                            start = max(0, i - 2)
                            end = min(len(words), i + 3)
                            suggestion = ' '.join(words[start:end])
                            if len(suggestion) > len(query) and len(suggestion) < 100:
                                suggestions.add(suggestion)
            
            return list(suggestions)[:limit]
            
        except Exception as e:
            self.logger.error(f"Failed to get search suggestions: {e}")
            return []
    
    def get_search_analytics(self, user, search_params: Dict[str, Any]) -> Dict[str, Any]:
        """
        Get search analytics and insights.
        
        Args:
            user: Current user for permission filtering
            search_params: Search parameters
            
        Returns:
            Dictionary of analytics data
        """
        try:
            # Get base queryset
            queryset = self._get_base_queryset(user)
            
            # Apply search filters
            queryset = self._apply_search_filters(queryset, search_params)
            
            # Calculate analytics
            total_referrals = queryset.count()
            
            # Status distribution
            status_distribution = queryset.values('status').annotate(count=Count('id'))
            
            # Priority distribution
            priority_distribution = queryset.values('priority').annotate(count=Count('id'))
            
            # Service type distribution
            service_type_distribution = queryset.values('service_type').annotate(count=Count('id'))
            
            # Average processing time
            avg_processing_time = queryset.filter(
                submitted_at__isnull=False,
                completed_at__isnull=False
            ).aggregate(
                avg_time=Avg(F('completed_at') - F('submitted_at'))
            )['avg_time']
            
            # Referral volume over time (last 30 days)
            thirty_days_ago = timezone.now() - timedelta(days=30)
            daily_volume = queryset.filter(
                created_at__gte=thirty_days_ago
            ).extra(
                select={'day': 'date(created_at)'}
            ).values('day').annotate(count=Count('id')).order_by('day')
            
            # Top specialisms
            specialism_counts = {}
            for referral in queryset.filter(required_specialisms__isnull=False):
                for specialism in referral.required_specialisms:
                    specialism_counts[specialism] = specialism_counts.get(specialism, 0) + 1
            
            top_specialisms = sorted(specialism_counts.items(), key=lambda x: x[1], reverse=True)[:10]
            
            return {
                'total_referrals': total_referrals,
                'status_distribution': list(status_distribution),
                'priority_distribution': list(priority_distribution),
                'service_type_distribution': list(service_type_distribution),
                'avg_processing_time': avg_processing_time.total_seconds() / 3600 if avg_processing_time else None,  # Convert to hours
                'daily_volume': list(daily_volume),
                'top_specialisms': top_specialisms,
            }
            
        except Exception as e:
            self.logger.error(f"Failed to get search analytics: {e}")
            return {}


class BulkOperationsService:
    """
    Service for bulk operations on referrals.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    def bulk_update_status(self, user, referral_ids: List[int], new_status: str, notes: str = "") -> Dict[str, Any]:
        """
        Bulk update status of referrals.
        
        Args:
            user: Current user
            referral_ids: List of referral IDs to update
            new_status: New status to set
            notes: Optional notes about the update
            
        Returns:
            Dictionary with results
        """
        try:
            # Get referrals that user has permission to update
            queryset = self._get_editable_queryset(user)
            referrals = queryset.filter(id__in=referral_ids)
            
            updated_count = 0
            errors = []
            
            for referral in referrals:
                try:
                    old_status = referral.status
                    referral.status = new_status
                    
                    # Update timestamp based on status
                    now = timezone.now()
                    if new_status == Referral.Status.SUBMITTED and not referral.submitted_at:
                        referral.submitted_at = now
                    elif new_status == Referral.Status.COMPLETED and not referral.completed_at:
                        referral.completed_at = now
                    
                    referral.save()
                    
                    # Log the change
                    self.logger.info(f"Bulk status update: Referral {referral.id} {old_status} -> {new_status}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Referral {referral.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(referral_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk status update failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(referral_ids),
            }
    
    def bulk_assign_referrer(self, user, referral_ids: List[int], new_referrer_id: int) -> Dict[str, Any]:
        """
        Bulk assign referrals to a new referrer.
        
        Args:
            user: Current user (must be admin)
            referral_ids: List of referral IDs to update
            new_referrer_id: ID of new referrer
            
        Returns:
            Dictionary with results
        """
        try:
            if not user.is_admin:
                return {
                    'success': False,
                    'error': 'Only administrators can reassign referrals',
                    'updated_count': 0,
                    'total_requested': len(referral_ids),
                }
            
            # Get new referrer
            from accounts.models import User
            try:
                new_referrer = User.objects.get(id=new_referrer_id, user_type=User.UserType.GP)
            except User.DoesNotExist:
                return {
                    'success': False,
                    'error': 'Invalid referrer ID',
                    'updated_count': 0,
                    'total_requested': len(referral_ids),
                }
            
            # Get referrals
            referrals = Referral.objects.filter(id__in=referral_ids)
            
            updated_count = 0
            errors = []
            
            for referral in referrals:
                try:
                    old_referrer = referral.referrer
                    referral.referrer = new_referrer
                    referral.save()
                    
                    self.logger.info(f"Bulk referrer assignment: Referral {referral.id} {old_referrer} -> {new_referrer}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Referral {referral.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(referral_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk referrer assignment failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(referral_ids),
            }
    
    def bulk_export(self, user, referral_ids: List[int], format: str = 'csv') -> Dict[str, Any]:
        """
        Bulk export referrals.
        
        Args:
            user: Current user
            referral_ids: List of referral IDs to export
            format: Export format ('csv', 'json', 'xlsx')
            
        Returns:
            Dictionary with export data
        """
        try:
            # Get referrals that user has permission to view
            queryset = self._get_viewable_queryset(user)
            referrals = queryset.filter(id__in=referral_ids).select_related('patient', 'referrer')
            
            if format == 'csv':
                return self._export_csv(referrals)
            elif format == 'json':
                return self._export_json(referrals)
            elif format == 'xlsx':
                return self._export_xlsx(referrals)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format}',
                }
            
        except Exception as e:
            self.logger.error(f"Bulk export failed: {e}")
            return {
                'success': False,
                'error': str(e),
            }
    
    def _get_editable_queryset(self, user):
        """Get queryset of referrals user can edit."""
        if user.is_gp:
            return Referral.objects.filter(referrer=user)
        elif user.is_admin:
            return Referral.objects.all()
        else:
            return Referral.objects.none()
    
    def _get_viewable_queryset(self, user):
        """Get queryset of referrals user can view."""
        if user.is_gp:
            return Referral.objects.filter(referrer=user)
        elif user.is_patient:
            return Referral.objects.filter(patient=user)
        elif user.is_admin:
            return Referral.objects.all()
        else:
            return Referral.objects.none()
    
    def _export_csv(self, referrals):
        """Export referrals as CSV."""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Referral ID', 'Patient Name', 'Referrer Name', 'Status', 'Priority',
            'Service Type', 'Modality', 'Presenting Problem', 'Created At', 'Submitted At'
        ])
        
        # Write data
        for referral in referrals:
            writer.writerow([
                referral.referral_id,
                f"{referral.patient.first_name} {referral.patient.last_name}",
                f"{referral.referrer.first_name} {referral.referrer.last_name}",
                referral.get_status_display(),
                referral.get_priority_display(),
                referral.get_service_type_display(),
                referral.get_modality_display(),
                referral.presenting_problem[:100] + '...' if len(referral.presenting_problem) > 100 else referral.presenting_problem,
                referral.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                referral.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if referral.submitted_at else '',
            ])
        
        return {
            'success': True,
            'data': output.getvalue(),
            'content_type': 'text/csv',
            'filename': f'referrals_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv',
        }
    
    def _export_json(self, referrals):
        """Export referrals as JSON."""
        import json
        
        data = []
        for referral in referrals:
            data.append({
                'referral_id': referral.referral_id,
                'patient_name': f"{referral.patient.first_name} {referral.patient.last_name}",
                'referrer_name': f"{referral.referrer.first_name} {referral.referrer.last_name}",
                'status': referral.status,
                'status_display': referral.get_status_display(),
                'priority': referral.priority,
                'priority_display': referral.get_priority_display(),
                'service_type': referral.service_type,
                'service_type_display': referral.get_service_type_display(),
                'modality': referral.modality,
                'modality_display': referral.get_modality_display(),
                'presenting_problem': referral.presenting_problem,
                'condition_description': referral.condition_description,
                'clinical_notes': referral.clinical_notes,
                'created_at': referral.created_at.isoformat(),
                'submitted_at': referral.submitted_at.isoformat() if referral.submitted_at else None,
                'completed_at': referral.completed_at.isoformat() if referral.completed_at else None,
            })
        
        return {
            'success': True,
            'data': json.dumps(data, indent=2),
            'content_type': 'application/json',
            'filename': f'referrals_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json',
        }
    
    def _export_xlsx(self, referrals):
        """Export referrals as XLSX."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Referrals"
            
            # Define headers
            headers = [
                'Referral ID', 'Patient Name', 'Referrer Name', 'Status', 'Priority',
                'Service Type', 'Modality', 'Presenting Problem', 'Condition Description',
                'Clinical Notes', 'Created At', 'Submitted At', 'Completed At'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row, referral in enumerate(referrals, 2):
                worksheet.cell(row=row, column=1, value=referral.referral_id)
                worksheet.cell(row=row, column=2, value=f"{referral.patient.first_name} {referral.patient.last_name}")
                worksheet.cell(row=row, column=3, value=f"{referral.referrer.first_name} {referral.referrer.last_name}")
                worksheet.cell(row=row, column=4, value=referral.get_status_display())
                worksheet.cell(row=row, column=5, value=referral.get_priority_display())
                worksheet.cell(row=row, column=6, value=referral.get_service_type_display())
                worksheet.cell(row=row, column=7, value=referral.get_modality_display())
                worksheet.cell(row=row, column=8, value=referral.presenting_problem)
                worksheet.cell(row=row, column=9, value=referral.condition_description)
                worksheet.cell(row=row, column=10, value=referral.clinical_notes)
                worksheet.cell(row=row, column=11, value=referral.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                worksheet.cell(row=row, column=12, value=referral.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if referral.submitted_at else '')
                worksheet.cell(row=row, column=13, value=referral.completed_at.strftime('%Y-%m-%d %H:%M:%S') if referral.completed_at else '')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {
                'success': True,
                'data': output.getvalue(),
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'filename': f'referrals_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            }
            
        except ImportError:
            return {
                'success': False,
                'error': 'openpyxl is required for XLSX export. Install with: pip install openpyxl',
            }
