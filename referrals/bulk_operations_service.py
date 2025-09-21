"""
Bulk operations service for referrals and appointments.
"""
import logging
from typing import Dict, List, Any, Optional
from django.db.models import Q, F
from django.utils import timezone
from datetime import datetime, timedelta
from .models import Referral, Candidate, Appointment, Message, Task


class AppointmentBulkOperationsService:
    """
    Service for bulk operations on appointments.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    def bulk_update_status(self, user, appointment_ids: List[int], new_status: str, notes: str = "") -> Dict[str, Any]:
        """
        Bulk update status of appointments.
        
        Args:
            user: Current user
            appointment_ids: List of appointment IDs to update
            new_status: New status to set
            notes: Optional notes about the update
            
        Returns:
            Dictionary with results
        """
        try:
            # Get appointments that user has permission to update
            queryset = self._get_editable_queryset(user)
            appointments = queryset.filter(id__in=appointment_ids)
            
            updated_count = 0
            errors = []
            
            for appointment in appointments:
                try:
                    old_status = appointment.status
                    appointment.status = new_status
                    
                    # Update timestamp based on status
                    now = timezone.now()
                    if new_status == Appointment.Status.CONFIRMED and not appointment.confirmed_at:
                        appointment.confirmed_at = now
                    elif new_status == Appointment.Status.COMPLETED and not appointment.completed_at:
                        appointment.completed_at = now
                    
                    # Add notes if provided
                    if notes:
                        appointment.notes = f"{appointment.notes}\n\nBulk update: {notes}".strip()
                    
                    appointment.save()
                    
                    # Log the change
                    self.logger.info(f"Bulk status update: Appointment {appointment.id} {old_status} -> {new_status}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Appointment {appointment.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(appointment_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk appointment status update failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(appointment_ids),
            }
    
    def bulk_reschedule(self, user, appointment_ids: List[int], new_datetime: str, notes: str = "") -> Dict[str, Any]:
        """
        Bulk reschedule appointments.
        
        Args:
            user: Current user
            appointment_ids: List of appointment IDs to reschedule
            new_datetime: New datetime in ISO format
            notes: Optional notes about the reschedule
            
        Returns:
            Dictionary with results
        """
        try:
            # Parse new datetime
            new_dt = timezone.datetime.fromisoformat(new_datetime.replace('Z', '+00:00'))
            
            # Get appointments that user has permission to update
            queryset = self._get_editable_queryset(user)
            appointments = queryset.filter(id__in=appointment_ids)
            
            updated_count = 0
            errors = []
            
            for appointment in appointments:
                try:
                    old_datetime = appointment.scheduled_at
                    appointment.scheduled_at = new_dt
                    
                    # Add notes if provided
                    if notes:
                        appointment.notes = f"{appointment.notes}\n\nBulk reschedule: {notes}".strip()
                    
                    appointment.save()
                    
                    # Log the change
                    self.logger.info(f"Bulk reschedule: Appointment {appointment.id} {old_datetime} -> {new_dt}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Appointment {appointment.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(appointment_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk reschedule failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(appointment_ids),
            }
    
    def bulk_assign_psychologist(self, user, appointment_ids: List[int], new_psychologist_id: int) -> Dict[str, Any]:
        """
        Bulk assign appointments to a new psychologist.
        
        Args:
            user: Current user (must be admin)
            appointment_ids: List of appointment IDs to update
            new_psychologist_id: ID of new psychologist
            
        Returns:
            Dictionary with results
        """
        try:
            if not user.is_admin:
                return {
                    'success': False,
                    'error': 'Only administrators can reassign appointments',
                    'updated_count': 0,
                    'total_requested': len(appointment_ids),
                }
            
            # Get new psychologist
            from accounts.models import User
            try:
                new_psychologist = User.objects.get(id=new_psychologist_id, user_type=User.UserType.PSYCHOLOGIST)
            except User.DoesNotExist:
                return {
                    'success': False,
                    'error': 'Invalid psychologist ID',
                    'updated_count': 0,
                    'total_requested': len(appointment_ids),
                }
            
            # Get appointments
            appointments = Appointment.objects.filter(id__in=appointment_ids)
            
            updated_count = 0
            errors = []
            
            for appointment in appointments:
                try:
                    old_psychologist = appointment.psychologist
                    appointment.psychologist = new_psychologist
                    appointment.save()
                    
                    self.logger.info(f"Bulk psychologist assignment: Appointment {appointment.id} {old_psychologist} -> {new_psychologist}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Appointment {appointment.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(appointment_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk psychologist assignment failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(appointment_ids),
            }
    
    def bulk_export(self, user, appointment_ids: List[int], format: str = 'csv') -> Dict[str, Any]:
        """
        Bulk export appointments.
        
        Args:
            user: Current user
            appointment_ids: List of appointment IDs to export
            format: Export format ('csv', 'json', 'xlsx')
            
        Returns:
            Dictionary with export data
        """
        try:
            # Get appointments that user has permission to view
            queryset = self._get_viewable_queryset(user)
            appointments = queryset.filter(id__in=appointment_ids).select_related('patient', 'psychologist', 'referral')
            
            if format == 'csv':
                return self._export_csv(appointments)
            elif format == 'json':
                return self._export_json(appointments)
            elif format == 'xlsx':
                return self._export_xlsx(appointments)
            else:
                return {
                    'success': False,
                    'error': f'Unsupported format: {format}',
                }
            
        except Exception as e:
            self.logger.error(f"Bulk appointment export failed: {e}")
            return {
                'success': False,
                'error': str(e),
            }
    
    def _get_editable_queryset(self, user):
        """Get queryset of appointments user can edit."""
        if user.is_gp:
            return Appointment.objects.filter(referral__referrer=user)
        elif user.is_patient:
            return Appointment.objects.filter(patient=user)
        elif user.is_psychologist:
            return Appointment.objects.filter(psychologist=user)
        elif user.is_admin:
            return Appointment.objects.all()
        else:
            return Appointment.objects.none()
    
    def _get_viewable_queryset(self, user):
        """Get queryset of appointments user can view."""
        if user.is_gp:
            return Appointment.objects.filter(referral__referrer=user)
        elif user.is_patient:
            return Appointment.objects.filter(patient=user)
        elif user.is_psychologist:
            return Appointment.objects.filter(psychologist=user)
        elif user.is_admin:
            return Appointment.objects.all()
        else:
            return Appointment.objects.none()
    
    def _export_csv(self, appointments):
        """Export appointments as CSV."""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Appointment ID', 'Referral ID', 'Patient Name', 'Psychologist Name', 
            'Scheduled At', 'Duration (min)', 'Status', 'Modality', 'Location', 'Notes'
        ])
        
        # Write data
        for appointment in appointments:
            writer.writerow([
                str(appointment.id),
                appointment.referral.referral_id,
                f"{appointment.patient.first_name} {appointment.patient.last_name}",
                f"{appointment.psychologist.first_name} {appointment.psychologist.last_name}",
                appointment.scheduled_at.strftime('%Y-%m-%d %H:%M:%S'),
                appointment.duration_minutes,
                appointment.get_status_display(),
                appointment.get_modality_display(),
                appointment.location,
                appointment.notes[:100] + '...' if len(appointment.notes) > 100 else appointment.notes,
            ])
        
        return {
            'success': True,
            'data': output.getvalue(),
            'content_type': 'text/csv',
            'filename': f'appointments_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv',
        }
    
    def _export_json(self, appointments):
        """Export appointments as JSON."""
        import json
        
        data = []
        for appointment in appointments:
            data.append({
                'appointment_id': str(appointment.id),
                'referral_id': appointment.referral.referral_id,
                'patient_name': f"{appointment.patient.first_name} {appointment.patient.last_name}",
                'psychologist_name': f"{appointment.psychologist.first_name} {appointment.psychologist.last_name}",
                'scheduled_at': appointment.scheduled_at.isoformat(),
                'duration_minutes': appointment.duration_minutes,
                'status': appointment.status,
                'status_display': appointment.get_status_display(),
                'modality': appointment.modality,
                'modality_display': appointment.get_modality_display(),
                'location': appointment.location,
                'notes': appointment.notes,
                'outcome_notes': appointment.outcome_notes,
                'created_at': appointment.created_at.isoformat(),
                'confirmed_at': appointment.confirmed_at.isoformat() if appointment.confirmed_at else None,
                'completed_at': appointment.completed_at.isoformat() if appointment.completed_at else None,
            })
        
        return {
            'success': True,
            'data': json.dumps(data, indent=2),
            'content_type': 'application/json',
            'filename': f'appointments_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json',
        }
    
    def _export_xlsx(self, appointments):
        """Export appointments as XLSX."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Appointments"
            
            # Define headers
            headers = [
                'Appointment ID', 'Referral ID', 'Patient Name', 'Psychologist Name',
                'Scheduled At', 'Duration (min)', 'Status', 'Modality', 'Location',
                'Notes', 'Outcome Notes', 'Created At', 'Confirmed At', 'Completed At'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row, appointment in enumerate(appointments, 2):
                worksheet.cell(row=row, column=1, value=str(appointment.id))
                worksheet.cell(row=row, column=2, value=appointment.referral.referral_id)
                worksheet.cell(row=row, column=3, value=f"{appointment.patient.first_name} {appointment.patient.last_name}")
                worksheet.cell(row=row, column=4, value=f"{appointment.psychologist.first_name} {appointment.psychologist.last_name}")
                worksheet.cell(row=row, column=5, value=appointment.scheduled_at.strftime('%Y-%m-%d %H:%M:%S'))
                worksheet.cell(row=row, column=6, value=appointment.duration_minutes)
                worksheet.cell(row=row, column=7, value=appointment.get_status_display())
                worksheet.cell(row=row, column=8, value=appointment.get_modality_display())
                worksheet.cell(row=row, column=9, value=appointment.location)
                worksheet.cell(row=row, column=10, value=appointment.notes)
                worksheet.cell(row=row, column=11, value=appointment.outcome_notes)
                worksheet.cell(row=row, column=12, value=appointment.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                worksheet.cell(row=row, column=13, value=appointment.confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if appointment.confirmed_at else '')
                worksheet.cell(row=row, column=14, value=appointment.completed_at.strftime('%Y-%m-%d %H:%M:%S') if appointment.completed_at else '')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return {
                'success': True,
                'data': output.getvalue(),
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'filename': f'appointments_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            }
            
        except ImportError:
            return {
                'success': False,
                'error': 'openpyxl is required for XLSX export. Install with: pip install openpyxl',
            }


class TaskBulkOperationsService:
    """
    Service for bulk operations on tasks.
    """
    
    def __init__(self):
        self.logger = logging.getLogger(f"{__name__}.{self.__class__.__name__}")
    
    def bulk_update_status(self, user, task_ids: List[int], new_status: str, notes: str = "") -> Dict[str, Any]:
        """
        Bulk update status of tasks.
        
        Args:
            user: Current user
            task_ids: List of task IDs to update
            new_status: New status to set
            notes: Optional notes about the update
            
        Returns:
            Dictionary with results
        """
        try:
            # Get tasks that user has permission to update
            queryset = self._get_editable_queryset(user)
            tasks = queryset.filter(id__in=task_ids)
            
            updated_count = 0
            errors = []
            
            for task in tasks:
                try:
                    old_status = task.is_completed
                    task.is_completed = new_status.lower() == 'completed'
                    
                    # Add notes if provided
                    if notes:
                        task.notes = f"{task.notes}\n\nBulk update: {notes}".strip()
                    
                    task.save()
                    
                    # Log the change
                    self.logger.info(f"Bulk status update: Task {task.id} {old_status} -> {task.is_completed}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Task {task.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(task_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk task status update failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(task_ids),
            }
    
    def bulk_assign_user(self, user, task_ids: List[int], new_user_id: int) -> Dict[str, Any]:
        """
        Bulk assign tasks to a new user.
        
        Args:
            user: Current user (must be admin)
            task_ids: List of task IDs to update
            new_user_id: ID of new user
            
        Returns:
            Dictionary with results
        """
        try:
            if not user.is_admin:
                return {
                    'success': False,
                    'error': 'Only administrators can reassign tasks',
                    'updated_count': 0,
                    'total_requested': len(task_ids),
                }
            
            # Get new user
            from accounts.models import User
            try:
                new_user = User.objects.get(id=new_user_id)
            except User.DoesNotExist:
                return {
                    'success': False,
                    'error': 'Invalid user ID',
                    'updated_count': 0,
                    'total_requested': len(task_ids),
                }
            
            # Get tasks
            tasks = Task.objects.filter(id__in=task_ids)
            
            updated_count = 0
            errors = []
            
            for task in tasks:
                try:
                    old_user = task.assigned_to
                    task.assigned_to = new_user
                    task.save()
                    
                    self.logger.info(f"Bulk user assignment: Task {task.id} {old_user} -> {new_user}")
                    
                    updated_count += 1
                    
                except Exception as e:
                    errors.append(f"Task {task.id}: {str(e)}")
            
            return {
                'success': True,
                'updated_count': updated_count,
                'total_requested': len(task_ids),
                'errors': errors,
            }
            
        except Exception as e:
            self.logger.error(f"Bulk user assignment failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'updated_count': 0,
                'total_requested': len(task_ids),
            }
    
    def _get_editable_queryset(self, user):
        """Get queryset of tasks user can edit."""
        if user.is_admin:
            return Task.objects.all()
        else:
            return Task.objects.filter(assigned_to=user)
    
    def _get_viewable_queryset(self, user):
        """Get queryset of tasks user can view."""
        if user.is_admin:
            return Task.objects.all()
        else:
            return Task.objects.filter(assigned_to=user)
